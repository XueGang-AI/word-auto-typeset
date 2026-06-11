"""
Batch processing service: manages multi-file typesetting jobs
with progress tracking, ZIP generation, and report creation.
"""

from __future__ import annotations

import json
import logging
import shutil
import tempfile
import threading
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.config import MAX_CONCURRENT_TASKS, OUTPUT_DIR
from app.schemas.models import (
    BatchJob,
    BatchProgress,
    BatchTask,
    ContentDocument,
    TaskStatus,
    TemplateConfig,
)
from app.services.content_parser import parse_content
from app.services.ai_recognizer import recognize_structure
from app.services.renderer import render_document
from app.services.template_service import build_style_brief
from app.utils.file_utils import safe_filename

logger = logging.getLogger(__name__)

# In-memory job store (upgrade to DB for production)
_jobs: dict[str, BatchJob] = {}
_lock = threading.Lock()


def create_batch_job(
    filenames: list[str],
    file_data_list: list[bytes],
    template_id: str,
    template_config: TemplateConfig,
    template_docx_path: Path | None = None,
) -> BatchJob:
    """
    Create a batch job from uploaded files.

    Args:
        filenames: Original filenames
        file_data_list: Raw file bytes for each file
        template_id: Template to use
        template_config: Loaded template config
        template_docx_path: Path to the template .docx file

    Returns:
        BatchJob with tasks enqueued
    """
    batch_id = f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    batch_dir = OUTPUT_DIR / batch_id
    batch_dir.mkdir(parents=True, exist_ok=True)
    (batch_dir / "input").mkdir(exist_ok=True)
    (batch_dir / "output").mkdir(exist_ok=True)
    (batch_dir / "json").mkdir(exist_ok=True)
    (batch_dir / "logs").mkdir(exist_ok=True)
    (batch_dir / "failed").mkdir(exist_ok=True)

    tasks = []
    for i, (fname, fdata) in enumerate(zip(filenames, file_data_list)):
        task_id = f"task_{batch_id}_{i:04d}"
        safe_name = safe_filename(fname)
        input_path = batch_dir / "input" / safe_name
        input_path.write_bytes(fdata)

        tasks.append(BatchTask(
            task_id=task_id,
            batch_id=batch_id,
            original_filename=fname,
            status=TaskStatus.pending,
        ))

    job = BatchJob(
        batch_id=batch_id,
        template_id=template_id,
        tasks=tasks,
        total_count=len(tasks),
    )

    with _lock:
        _jobs[batch_id] = job

    # Start processing in background
    threading.Thread(
        target=_process_batch,
        args=(batch_id, template_config, template_docx_path),
        daemon=True,
    ).start()

    return job


def get_batch_progress(batch_id: str) -> BatchProgress | None:
    """Get the current progress of a batch job."""
    job = _jobs.get(batch_id)
    if not job:
        return None

    return BatchProgress(
        batch_id=job.batch_id,
        status=job.status,
        total=job.total_count,
        completed=job.success_count + job.failed_count,
        success=job.success_count,
        failed=job.failed_count,
        low_confidence=job.low_confidence_count,
    )


def get_batch_job(batch_id: str) -> BatchJob | None:
    """Get the full batch job status."""
    return _jobs.get(batch_id)


def _process_batch(
    batch_id: str,
    template_config: TemplateConfig,
    template_docx_path: Path | None = None,
) -> None:
    """
    Process all tasks in a batch job concurrently.

    For each file:
    1. Parse content → paragraphs list
    2. AI recognize → ContentDocument (content.json)
    3. Render → formatted .docx
    4. Log results
    """
    job = _jobs.get(batch_id)
    if not job:
        return

    job.status = TaskStatus.processing
    batch_dir = OUTPUT_DIR / batch_id

    with ThreadPoolExecutor(max_workers=MAX_CONCURRENT_TASKS) as executor:
        futures = {}
        for task in job.tasks:
            input_path = batch_dir / "input" / safe_filename(task.original_filename)
            if not input_path.exists():
                task.status = TaskStatus.failed
                task.error_message = "输入文件不存在"
                job.failed_count += 1
                continue

            future = executor.submit(
                _process_single_file,
                task,
                input_path,
                batch_dir,
                template_config,
                template_docx_path,
            )
            futures[future] = task

        for future in as_completed(futures):
            task = futures[future]
            try:
                future.result()
            except Exception as e:
                task.status = TaskStatus.failed
                task.error_message = str(e)
                job.failed_count += 1
                logger.error(f"Task {task.task_id} failed: {e}")

    # Finalize
    job.status = TaskStatus.completed
    _finalize_batch(batch_id)


def _process_single_file(
    task: BatchTask,
    input_path: Path,
    batch_dir: Path,
    template_config: TemplateConfig,
    template_docx_path: Path | None = None,
) -> None:
    """Process a single file through the pipeline."""
    task.status = TaskStatus.processing
    log_entries: list[str] = []
    warnings: list[str] = []

    try:
        safe_name = safe_filename(task.original_filename)
        stem = Path(safe_name).stem

        # ── Step 1: Parse content ──
        log_entries.append(f"[{datetime.now().isoformat()}] 开始解析: {task.original_filename}")
        paragraphs = parse_content(input_path)
        log_entries.append(f"[{datetime.now().isoformat()}] 解析完成: {len(paragraphs)} 个段落")

        if not paragraphs:
            raise ValueError("内容文档为空，无法处理")

        # ── Step 2: AI structure recognition ──
        log_entries.append(f"[{datetime.now().isoformat()}] 开始AI结构识别")
        style_brief = build_style_brief(template_config)
        content_doc = recognize_structure(
            paragraphs, task.original_filename, style_brief=style_brief,
        )
        log_entries.append(f"[{datetime.now().isoformat()}] AI识别完成")

        # Save content.json
        content_json_path = batch_dir / "json" / f"{stem}_content.json"
        content_json_path.write_text(
            content_doc.model_dump_json(indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        task.content_id = content_doc.content_id

        # Check warnings
        if content_doc.warnings:
            warnings.extend(content_doc.warnings)
            log_entries.extend(f"[WARN] {w}" for w in content_doc.warnings)

        if content_doc.has_low_confidence:
            task.has_low_confidence = True
            log_entries.append("[WARN] 存在低置信度识别结果")

        # ── Step 3: Check for critical issues ──
        has_body = any(
            p.para_type.value == "body_text"
            for article in content_doc.articles
            for p in article.paragraphs
        )
        if not has_body:
            raise ValueError("未检测到正文内容")

        # ── Step 4: Render ──
        output_path = batch_dir / "output" / f"{stem}_排版后.docx"
        log_entries.append(f"[{datetime.now().isoformat()}] 开始渲染")
        render_document(content_doc, template_config, output_path, template_docx_path)
        log_entries.append(f"[{datetime.now().isoformat()}] 渲染完成")

        task.output_filename = f"{stem}_排版后.docx"

        # ── Step 5: Mark success ──
        task.status = TaskStatus.completed
        task.warnings = warnings
        log_entries.append(f"[{datetime.now().isoformat()}] 处理成功")

    except Exception as e:
        task.status = TaskStatus.failed
        task.error_message = str(e)
        log_entries.append(f"[{datetime.now().isoformat()}] 处理失败: {e}")

        # Move input file to failed directory
        failed_path = batch_dir / "failed" / input_path.name
        try:
            shutil.copy2(input_path, failed_path)
        except Exception:
            pass

    finally:
        # Save log
        log_path = batch_dir / "logs" / f"{safe_name}_log.json"
        log_path.write_text(
            json.dumps({
                "task_id": task.task_id,
                "original_filename": task.original_filename,
                "status": task.status.value,
                "error": task.error_message,
                "warnings": warnings,
                "log": log_entries,
            }, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )


def _finalize_batch(batch_id: str) -> None:
    """Create ZIP and report after all tasks complete."""
    job = _jobs.get(batch_id)
    if not job:
        return

    batch_dir = OUTPUT_DIR / batch_id
    job.success_count = sum(1 for t in job.tasks if t.status == TaskStatus.completed)
    job.failed_count = sum(1 for t in job.tasks if t.status == TaskStatus.failed)
    job.low_confidence_count = sum(1 for t in job.tasks if t.has_low_confidence)

    # ── Create ZIP ──
    zip_path = batch_dir / "result.zip"
    output_dir = batch_dir / "output"
    if output_dir.exists():
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for f in sorted(output_dir.iterdir()):
                if f.is_file():
                    zf.write(f, arcname=f.name)
    job.zip_download_url = f"/api/typeset/batch/{batch_id}/download/zip"

    # ── Create report.xlsx ──
    report_path = _create_report(job, batch_dir)
    job.report_download_url = f"/api/typeset/batch/{batch_id}/download/report"

    logger.info(f"Batch {batch_id} completed: {job.success_count}/{job.total_count}")


def _create_report(job: BatchJob, batch_dir: Path) -> Path:
    """Create an Excel report for the batch job."""
    wb = Workbook()
    ws = wb.active
    ws.title = "处理报告"

    # Styles
    header_font = Font(name="微软雅黑", bold=True, size=12)
    header_fill = PatternFill(start_color="8B5E34", end_color="8B5E34", fill_type="solid")
    header_font_white = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
    success_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fail_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

    # Summary sheet
    ws.append(["Word 自动套版 - 批量处理报告"])
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14)

    ws.append([])
    ws.append(["批次ID", job.batch_id])
    ws.append(["模板ID", job.template_id])
    ws.append(["处理时间", datetime.now().isoformat()])
    ws.append(["总文件数", job.total_count])
    ws.append(["成功数", job.success_count])
    ws.append(["失败数", job.failed_count])
    ws.append(["低置信度数", job.low_confidence_count])
    ws.append(["成功率", f"{job.success_count / max(job.total_count, 1) * 100:.1f}%"])
    ws.append([])

    # Detail headers
    headers = ["文件名", "状态", "输出文件", "低置信度", "错误信息", "警告"]
    header_row = ws.max_row + 1
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Detail rows
    for task in job.tasks:
        row = [
            task.original_filename,
            task.status.value,
            task.output_filename or "",
            "是" if task.has_low_confidence else "否",
            task.error_message or "",
            "; ".join(task.warnings) if task.warnings else "",
        ]
        ws.append(row)
        row_num = ws.max_row
        if task.status == TaskStatus.completed:
            fill = warn_fill if task.has_low_confidence else success_fill
        else:
            fill = fail_fill
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).fill = fill

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 40

    report_path = batch_dir / "report.xlsx"
    wb.save(str(report_path))
    return report_path
